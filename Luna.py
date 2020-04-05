import speech_recognition as sr
from random import randint as random
import pyttsx3
from colorama import Fore, Back, Style,init
import openpyxl as xl


cur_sheet = None
wb = None
init()

def workbook(file = "friend.xlsx"):
    
    global cur_sheet,wb
    wb = xl.load_workbook(file)
    cur_sheet = wb.active
    return None

workbook()




def voice(message):

        engine = pyttsx3.init()
        engine.setProperty('rate',140)
        engine.setProperty('volume',1.0)
        voices = engine.getProperty('voices')
        engine.setProperty('voice',voices[1].id)
        engine.say(message)
        engine.runAndWait()
        engine.stop()
        return



def suggestion():

        suggestions = [cur_sheet.cell(random(1,cur_sheet.max_row),1).value for x in range(4)]

        main = init()

        print('\nYou May Ask : \n ')

        for suggest in range(len(suggestions)):

            print(f'{suggest+1})',Fore.GREEN+suggestions[suggest].capitalize(),sep= ' ')
            print(Style.RESET_ALL)

        print(Style.RESET_ALL)


def audio(calcus = None):

        r=sr.Recognizer()

        with sr.Microphone() as source:

            print("\nListening:\n")

            voice("listening")

            if calcus == None:

                suggestion()

            else:

            	pass

            audio=r.listen(source)

        try:

            text=r.recognize_google(audio)

        except:

            print('Luna: Sorry Voice could not recognized')

            voice("Sorry Voice could not recognized")

            return None

        else:
             
            print('\nMe:{}'.format(text.capitalize()))

            return text.upper()


def keyboard(calcus = None):

    try:
 
        if calcus == None:

            suggestion()

        else:

            pass

        command = input("\nType : ")


    except ValueError:

        print("\nLuna : Please enter words.\n")

        voice("Please enter words")

    return command.upper()



def arithmetic(expression):  # call with joined string

        try:

            expression = expression.replace(" ","")

            listed_expression = []

            exp_element = ""

            arr_expression = list(expression)

            if arr_expression[0] == "-":

                negative = "-"

                for element in arr_expression[1::]:

                    if element.isdigit():

                        negative += element

                        arr_expression.pop(1)

                    else:

                        break

                arr_expression[0] = negative

            for item in arr_expression:

                if item.isdigit() or len(item) >= 2:

                    exp_element += item

                else:

                    listed_expression.append(exp_element)
                    #print(listed_expression)


                    listed_expression.append(item)

                    exp_element = ""

            listed_expression.append(exp_element)

           # print(listed_expression)

            operator = listed_expression[1].upper()

            operand_one = int(listed_expression[0])

            operand_two = int(listed_expression[2])

            if operator == "+":
                sub_result = operand_one + operand_two
            elif operator == "X":
                sub_result = operand_one * operand_two
            elif operator == "/":
                sub_result = int(operand_one / operand_two)
            elif operator == "-":
                sub_result = operand_one - operand_two
            else:
                return "Sorry! give the input and 'Try Again'"            

            listed_expression.pop(0)
            listed_expression.pop(0)
            listed_expression.pop(0)
            listed_expression.insert(0,str(sub_result))
            if len(listed_expression) != 1:
                return arithmetic("".join(listed_expression))
            else:
                return int(listed_expression[0])

        
        except ValueError:

            if expression != "CLOSECALCULATOR":

                return "Please give the input as a integer."

        except IndexError:
            return "Please give atleast two number to calculate."

        except AttributeError:
            return "Please give atleast two number to calculate."




def introduction():

    print("\nLuna : Hello There! I'm Luna")
    voice("Hello There! I'm Luna")

    print("\nLuna : I am created to make a good conversation with others.\n")
    voice("I am created to make a good conversation with others")

    print("Luna : I can also help you with the calculation",Fore.GREEN+"(say 'switch to Calculator' to activate calculator)\n")
    voice("I can also help you with the calculation")

    print(Style.RESET_ALL)

    print("Luna : say 'bye' to terminate the program.\n")
    voice("say 'bye' to terminate the program")


    voice("\nenter your Name")

    user_name = input("Enter your Name : ")

    voice(f"Hello {user_name}. Do you wish to continue?")

    wish = input(f"\nHello {user_name}. Do you wish to continue? (y/n)\n")

    return wish.upper()



def database(user_question):

    workbook()

    analyis = []    

    present = False    

    for row in range(1,cur_sheet.max_row+1):    

        question = cur_sheet.cell(row,1).value
        
        answer = cur_sheet.cell(row,2).value    

        analyis.append([question,answer])    

    for sub_arr in range(len(analyis)):    

        if user_question == analyis[sub_arr][0]:    

            print(f"Luna : {analyis[sub_arr][1].capitalize()}")    

            voice(analyis[sub_arr][1])    

            present = True    

            break    

    if not present:    

        print(Fore.YELLOW+"\nSorry! This question is not in my data.\n")

        voice("sorry This question is not in my data")

        voice("Do you want to add in that")

        choice = input("\nDo you want to add in that? (y/n)\n")    

        if choice.upper() == 'Y':    

            voice("Please, enter the answer too")

            answer = input("\nPlease, enter the answer too! : \n")

            voice("Sure? Do you want to upload")

            choice_two = input("\nSure? Do you want to upload (y/n)\n")    

            if choice_two.upper() == 'Y':    

                cur_sheet.cell(cur_sheet.max_row+1,1).value = user_question  

                cur_sheet.cell(cur_sheet.max_row,2).value = answer.upper()    

                wb.save("friend.xlsx")
            else:
                pass
        else:
            pass    

        print(Style.RESET_ALL)    

    else:
        pass     


def conversation():

    user_wish = introduction()

    if user_wish != "Y":

        print("\nLuna : Thank you! Let's catch up Later.\n")

        voice("Thank you! Let's catch up Later")

    else:

        
        user_input_type_wish = ""

        voice("Do you wish to continue on Keyboard or Voice")

        while not (user_input_type_wish == "KEYBOARD" or user_input_type_wish == "VOICE" or user_input_type_wish == "BYE"):

            user_input_type_wish = input("\nDo you wish to continue on Keyboard or Voice?\n( Type 'Bye' to terminate the program ) : ").upper()

        main_loop = True

        Keyboard_type_input = voice_type_input = False

        while main_loop:

            if user_input_type_wish == "KEYBOARD":

                Keyboard_type_input = True

                voice_type_input = False

                message = keyboard()

                if message != 'VOICE':

                    pass

                else:

                    user_input_type_wish = "VOICE"

                    message = None

            elif user_input_type_wish == "VOICE":

                Keyboard_type_input = False

                voice_type_input = True

                message = audio()

                if message != 'KEYBOARD':

                    pass

                else:

                    user_input_type_wish = "KEYBOARD"

                    message = None

            elif user_input_type_wish == "BYE":

                print("\nLuna : Thank you! Let's catch up Later. Bye.")

                voice("Thank you! Let's catch up Later")

                voice("bye")

                main_loop = False

                break 


            if message == None or message == "":

            	pass

            elif message == "SWITCH TO CALCULATOR":

                print("\nLuna : Switched to Calculator\n")

                voice("Switched to CALCULATOR")

                cal_question = ""

                if voice_type_input:

                    print(f"Use this keywords.\n{1} ) Plus\n{2} ) Minos\n{3} )'Times' for multiply\n{4} ) Divide by\n'Close Calculator to terminate.\n'")

                    voice("use this keyword for your calculation")

                    cal_question = ""

                    while not(cal_question == "CLOSE CALCULATOR"):

                        cal_question = audio(calcus = "On")
                    
                        result = arithmetic(cal_question)

                        print("\n",result,"\n")

                        voice(result)

                    print("\nLuna : Calculator Closed\n")

                    voice("Calculator Closed")

                elif Keyboard_type_input:

                    print(f"Use this keywords.\n{1} ) +\n{2} ) -\n{3} )'x' for multiply\n{4} ) '/' for Divide\n'Close Calculator to terminate.\n'")

                    voice("Use this keywords for your calculation")

                    cal_question = ""

                    while not(cal_question == "CLOSE CALCULATOR"):

                        cal_question = keyboard(calcus = "on")
                        
                        result = arithmetic(cal_question)

                        print("\n",result,"\n")

                        voice(result)

                    print("\nLuna : Calculator Closed\n")

                    voice("Calculator Closed")

            elif message.startswith("BYE") or message.startswith("BHAI"):

                print("\nLuna : Thank you! Let's catch up Later. Bye.")

                voice("Thank you! Let's catch up Later")

                voice("bye")

                main_loop = False

            else:

            	database(message)

conversation()




            




























            













